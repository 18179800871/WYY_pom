'''
    基于:excel实现自动化
    既然关键字已经封装我可否通过对数据的读取,让他自己生成脚本

'''
import bold as bold
import openpyxl
import row as row
import self as self
from openpyxl.drawing import fill
from logs.log import Logger
from openpyxl.styles import PatternFill, Font

from class07.key_word_web.keyword_web import WebKeys
# 创建日志对象
log = Logger().get_logger()
excel = openpyxl.load_workbook('../data/xxx.xlsx')
sheet = excel['Sheet1']



log.info('获取{0}内容成功，现在开始执行自动化测试......'.format(sheet))
# 读取excel内容,实现文件驱动自动化执行
for value in sheet.values:
    # 定义一个字典参数,用于接受excel中所有参数内容
    # 基于A列进行判断是否为测试用例
    args = {}
    #定位方法
    args['name'] = value[2]
    #定位路径
    args['value'] = value[3]
    #输入内容
    args['txt'] = value[4]
    #预期结果
    args['expect'] = value[6]
    if type(value[0]) is int:
        '''
            在调用关键字时,分为几类情况
                1.关键字实例化
                2.断言类型的关键字
                    因为断言的关键字,需要获取元素,还需要对比
        '''
        #判断是否实例化
        if value[1] == 'open_browser':
                log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
                wk = WebKeys(value[4])

        # 断言关键字执行,首先执行断言,在判断断言是否成功
        elif 'assert' in value[1]:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            status = getattr(wk ,value[1])(**args)
            # 依据status来判定写入的内容是True 还是Failed还是pass
            if status:
                log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
                #写入当前行的结果值
                # sheet.cell(row='编号+1', column=8) #指定单元格
                sheet.cell(row=value[0]+1,column=8).value = 'Pass'
                sheet.cell(row=value[0]+1, column=8).fill = PatternFill('solid', fgColor='AACF91') # 色块
                sheet.cell(row=value[0]+1, column=8).font = Font(bold=True)  # 加粗
            else:
                log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
                sheet.cell(row=value[0]+1,column=8).value = 'Failed'
                sheet.cell(row=value[0] + 1, column=8).fill = PatternFill('solid', fgColor='FF0000')
                sheet.cell(row=value[0] + 1, column=8).font = Font(bold=True)  # 加粗
        # 非特殊关键字,通过反射机制实现
        else:
            log.info('现在执行关键字:{0}，操作描述：{1}'.format(value[1], value[5]))
            getattr(wk,value[1])(**args)
            # wk.open(**args)
            # 原有的open的函数是 def open(self,url)
            '''
                如果是open,需要传入value[4]
                如果是input,则需要传入,value[2],value[3],value[4]
                如果是click,则需要传入value[2],value[3]
                **不点长传参,不定义参数数量的多少,说白了就是一个字典
                但是字典是键值对形式.
                
            '''
excel.save('../data/xxx.xlsx')
excel.close()

















