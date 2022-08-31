'''
    excel的数据驱动实现
        1.Excel操作流程:先操作workbook,在操作sheet,在操作cell(value)
        2.对于模块的API建议在使用前百度一下
        3.excel操作结束后记得释放
        
修改excel格式
'''
#导入模块
import openpyxl
# excel 操作流程
#操作工作薄:指定文件路径,进行文件读取,类似于open()函数的文件读取操作
from openpyxl.styles import Font, PatternFill

excel = openpyxl.load_workbook('../data/xxx.xlsx')
print(excel)

# 获取sheet :基于工作簿来获取sheet页
names = excel.sheetnames
print(names)
print(type(names))
for name in names:
    print(name)
# 操作单元格 指定sheet页,再进行操作
sheet = excel['Sheet1']
print(sheet)
# 获取单个单元格内容
print(sheet.vlues)
for value in sheet.vlues:
    print(value[0])
# 指定单元格进行内容获取
value = sheet['A5'].value
print(value)

# 获取行: 最大行数
rows = sheet.max_row
print(rows)

# 获取列:最大列数
column = sheet.max_colum
print(column)
# 写入 :基于单元格导入
# 写入前一定确认操作文件处于关闭的情况下
sheet['H1'] = 'asd'
# 写入之后一定需要保存,保存对象是excel,而不是sheet或者cell
excel.save('../data/xxx.xlsx')
# 在所有操作结束后,记得释放
excel.close()
# 定义EXcel格式
class EXcelconf:
    def cell_write(self,value,sheet,row):
        bold = Font(bold=True)
        if value == 'pass':
            fill = PatternFill('solid',fgColor='AACF91')
        elif value =='false':
            fill  = PatternFill('solid',fgColor='FF0000')
        else:
            pass














